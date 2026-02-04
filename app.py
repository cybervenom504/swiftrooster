import streamlit as st
from calendar import monthrange
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import os

# ---------------- CONFIG ----------------
st.set_page_config(page_title="SwiftRoster Pro", layout="wide")
st.title("✈️ SwiftRoster Pro – Airline Duty Roster")

WORKERS_PER_DAY = 8
SUPERVISORS = ["SUPERVISOR A", "SUPERVISOR B", "SUPERVISOR C"]

# ---------------- STAFF ----------------
workers = [
    "ONYEWUENYI", "NDIMELE", "BELLO", "FASEYE",
    "IWUNZE", "OZUA", "JAMES", "OLABANJI",
    "NURUDEEN", "ENEH", "MUSA", "SANI",
    "ADENIJI", "JOSEPH", "IDOWU"
]

# ---------------- SIDEBAR ----------------
month = st.sidebar.selectbox("Month", range(1, 13), index=0)
year = st.sidebar.number_input("Year", value=2026)

st.sidebar.info("Leave format:\nNAME: 5, 10, 22")
leave_input = st.sidebar.text_area("Leave")

leave = {}
if leave_input:
    for line in leave_input.split("\n"):
        if ":" in line:
            n, d = line.split(":")
            leave[n.strip()] = [int(x) for x in d.split(",") if x.strip().isdigit()]

# ---------------- GENERATE ----------------
if st.button("🚀 Generate Roster (Excel + PDF)"):
    days = monthrange(year, month)[1]

    worker_count = {w: 0 for w in workers}
    supervisor_count = {s: 0 for s in SUPERVISORS}

    roster = {}
    supervisor_day = {}

    for d in range(1, days + 1):
        sup = sorted(SUPERVISORS, key=lambda x: supervisor_count[x])[0]
        supervisor_count[sup] += 1
        supervisor_day[d] = sup

        available = [w for w in workers if d not in leave.get(w, [])]
        available.sort(key=lambda x: worker_count[x])
        assigned = available[:WORKERS_PER_DAY]

        for w in assigned:
            worker_count[w] += 1

        roster[d] = assigned

    # ---------------- EXCEL ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "ROSTER"

    thick = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )
    center = Alignment(horizontal="center", vertical="center")
    bold = Font(bold=True)

    # Logo
    if os.path.exists("logo.png"):
        img = XLImage("logo.png")
        img.width = 150
        img.height = 70
        ws.add_image(img, "A1")

    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=days + 1)
    ws["C1"] = f"ROYAL AIR MAROC ROSTER FOR {datetime(year, month, 1).strftime('%B %Y').upper()}"
    ws["C1"].font = Font(size=14, bold=True)
    ws["C1"].alignment = center

    row = 4

    # DATE
    ws[f"A{row}"] = "DATE"
    ws[f"A{row}"].font = bold
    for d in range(1, days + 1):
        ws.cell(row=row, column=d + 1, value=d)
    row += 1

    # DAY
    ws[f"A{row}"] = "DAY"
    ws[f"A{row}"].font = bold
    for d in range(1, days + 1):
        ws.cell(row=row, column=d + 1, value=datetime(year, month, d).strftime("%a")[0])
    row += 1

    # SUPERVISOR
    ws[f"A{row}"] = "SUPERVISOR"
    ws[f"A{row}"].font = bold
    for d in range(1, days + 1):
        ws.cell(row=row, column=d + 1, value=supervisor_day[d])
    row += 1

    # Workers + M/A rotation
    for w in workers:
        ws[f"A{row}"] = w
        ws[f"A{row}"].font = bold
        duties = 0
        for d in range(1, days + 1):
            val = "X"
            if d in leave.get(w, []):
                val = "L"
            elif w in roster[d]:
                duties += 1
                val = "M" if duties % 2 else "A"
            ws.cell(row=row, column=d + 1, value=val)
        row += 1

    # Styling
    for r in ws.iter_rows(min_row=4, max_row=row - 1, min_col=1, max_col=days + 1):
        for c in r:
            c.border = thick
            c.alignment = center

    ws.column_dimensions["A"].width = 22
    for i in range(2, days + 2):
        ws.column_dimensions[get_column_letter(i)].width = 4

    # Footer
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=days + 1)
    ws[f"A{row}"] = (
        "RESUMPTION TIME: M 2330HRS / A 1200HRS   "
        "KEYNOTE: X-OFF, M-MORNING, A-AFTERNOON, L-LEAVE"
    )
    ws[f"A{row}"].alignment = center

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=days // 2)
    ws.merge_cells(start_row=row, start_column=days // 2 + 1, end_row=row, end_column=days + 1)
    ws[f"A{row}"] = "PREPARED BY: ____________________"
    ws[f"{get_column_letter(days // 2 + 1)}{row}"] = "CERTIFIED BY: ____________________"

    ws.protection.sheet = True
    ws.protection.enable()

    excel_file = f"ROYAL_AIR_MAROC_ROSTER_{month}_{year}.xlsx"
    wb.save(excel_file)

    # ---------------- PDF ----------------
    pdf_file = excel_file.replace(".xlsx", ".pdf")
    pdf = SimpleDocTemplate(pdf_file, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    table_data = [["NAME"] + [str(i) for i in range(1, days + 1)]]
    for w in workers:
        row_data = [w]
        for d in range(1, days + 1):
            if d in leave.get(w, []):
                row_data.append("L")
            elif w in roster[d]:
                row_data.append("M")
            else:
                row_data.append("X")
        table_data.append(row_data)

    table = Table(table_data)
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONT", (0,0), (-1,0), "Helvetica-Bold"),
    ]))

    pdf.build([
        Paragraph(f"<b>ROYAL AIR MAROC ROSTER – {month}/{year}</b>", styles["Title"]),
        table
    ])

    # ---------------- DOWNLOADS ----------------
    with open(excel_file, "rb") as f:
        st.download_button("📥 Download Excel", f, file_name=excel_file)

    with open(pdf_file, "rb") as f:
        st.download_button("📥 Download PDF", f, file_name=pdf_file)
